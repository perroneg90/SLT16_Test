import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '.python_packages', 'lib', 'site-packages'))

import logging
import azure.functions as func
import base64
import io
import pandas as pd
import openpyxl
from openpyxl.utils import range_boundaries
import datetime
import traceback
import json
import re
import pyodbc

# ---------------------------------------------------------------------------
#  Helpers
# ---------------------------------------------------------------------------

def format_text_for_excel(value: str) -> str:
    """Formatta il testo per Excel senza aggiungere l'apostrofo visibile."""
    v = str(value).strip()
    if v == '' or v.lower() in ('nan', 'none'):
        return ''
    return v  # Rimosso l'apostrofo


def clean_piva(value: str) -> str:
    """Rimuove il prefisso IT/it senza aggiungere l'apostrofo."""
    if pd.isna(value) or value is None or str(value).lower() in ('none', 'nan', ''):
        return ''
    v = str(value).strip()
    v = re.sub(r'^\s*IT', '', v, flags=re.IGNORECASE).strip()
    return v  # Rimosso l'apostrofo


def _clean_numeric(value):
    """Pulisce la stringa numerica da = , % ecc."""
    if pd.isna(value):
        return ''
    s = str(value).strip()
    if s.startswith('='):
        s = s[1:]
    return s.replace(',', '.').replace('%', '').strip()


def _to_float(value):
    s = _clean_numeric(value)
    return float(s) if re.match(r'^-?\d+(\.\d+)?$', s) else value


def _percent_times_100(value):
    f = _to_float(value)
    return f * 100 if isinstance(f, (int, float)) else f


def extract_info_from_cf(codice_fiscale, db_password="password_placeholder"):
    """Estrae sesso, data, comune, provincia da CF persona fisica."""
    try:
        cf = str(codice_fiscale).strip().upper()
        # Definisco la regex come variabile separata per evitare problemi di sintassi
        cf_pattern = r'^[A-Z]{6}\d{2}[A-Z]\d{2}[A-Z]\d{3}[A-Z]$'
        
        if not cf or len(cf) != 16 or not re.match(cf_pattern, cf):
            logging.debug(f"Codice fiscale non valido: {cf}")
            return None
            
        anno = int(cf[6:8])
        anno_corrente = datetime.datetime.now().year % 100
        secolo = 1900 if anno > anno_corrente else 2000
        anno_completo = secolo + anno
        mesi_dict = dict(zip("ABCDEHLMPRST", range(1, 13)))
        mese = mesi_dict.get(cf[8], 1)
        giorno = int(cf[9:11])
        sesso = 'F' if giorno > 40 else 'M'
        if sesso == 'F':
            giorno -= 40
        data_nascita = f"{giorno:02d}{mese:02d}{anno_completo}"
        codice_comune = cf[11:15]
        comune_nascita = provincia_nascita = ""
        
        # Connessione al database per recuperare i dati del comune
        try:
            # Prova con un solo driver, il più semplice, per verificare la connessione
            conn_string = "Driver={SQL Server};Server=euwdaitasksql02.database.windows.net;Database=EUWDAITASKSDB06;Uid=dbWSS;Pwd=" + db_password
            logging.info("Tentativo di connessione al DB")
            conn = pyodbc.connect(conn_string)
            cursor = conn.cursor()
            
            # Esegue query per recuperare comune e provincia dal codice catastale
            query = "SELECT Comune, Provincia FROM [dbo].[Comuni] WHERE CodiceCatastale = ?"
            cursor.execute(query, (codice_comune,))
            row = cursor.fetchone()
            
            if row:
                comune_nascita = row.Comune
                provincia_nascita = row.Provincia
            else:
                comune_nascita = f"Da Codice: {codice_comune}"
            
            cursor.close()
            conn.close()
        except Exception as db_err:
            logging.error(f"Errore connessione DB: {db_err}")
            comune_nascita = f"Da Codice: {codice_comune}"
        
        return {
            'sesso': sesso,
            'data_nascita': data_nascita,
            'comune_nascita': comune_nascita,
            'provincia_nascita': provincia_nascita
        }
    except Exception as e:
        logging.debug(f"Errore CF {codice_fiscale}: {e}")
        return None


def main(req: func.HttpRequest) -> func.HttpResponse:
    logging.info('Python HTTP trigger function processed a request.')

    try:
        req_body = req.get_json()
    except ValueError:
        return func.HttpResponse(
            json.dumps({"error": "Inserire il file Excel in formato base64"}),
            status_code=400,
            mimetype="application/json"
        )

    if not req_body or 'excelBase64' not in req_body:
        return func.HttpResponse(
            json.dumps({"error": "Inserire il file Excel in formato base64"}),
            status_code=400,
            mimetype="application/json"
        )

    # La password del DB verrà sostituita in fase di esecuzione
    db_password = "1e2A=L3~0IT~lgu*RDela*knWs~hp)Z1"
    if 'dbPassword' in req_body:
        db_password = req_body['dbPassword']

    try:
        excel_bytes = base64.b64decode(req_body['excelBase64'])
        excel_file = io.BytesIO(excel_bytes)

        # -------------------------------------------------------------------
        #  Lettura workbook (data_only=True)
        # -------------------------------------------------------------------
        table_data = None
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            for sh in wb.sheetnames:
                ws = wb[sh]
                if hasattr(ws, '_tables') and 'TableA2' in ws._tables:
                    tbl = ws._tables['TableA2']
                    min_c, min_r, max_c, max_r = range_boundaries(tbl.ref)
                    hdr = [ws.cell(min_r, c).value for c in range(min_c, max_c + 1)]
                    rows = [
                        [ws.cell(r, c).value for c in range(min_c, max_c + 1)]
                        for r in range(min_r + 1, max_r + 1)
                    ]
                    table_data = pd.DataFrame(rows, columns=hdr)
                    break
        except Exception as e:
            logging.warning(f"openpyxl: {e}")

        if table_data is None:
            logging.info("Fallback pandas.read_excel()")
            excel_file.seek(0)
            table_data = pd.read_excel(excel_file)

        if table_data is None or table_data.empty:
            return func.HttpResponse(
                json.dumps({"error": "Nessun dato trovato"}),
                status_code=404,
                mimetype="application/json"
            )

        # -------------------------------------------------------------------
        #  Normalizzazione colonne base
        # -------------------------------------------------------------------
        required_columns = [
            'Cognome / Denominazione Percipiente', 'Nome Percipiente',
            'P. Iva', 'C. Fisc', 'Causale CU', 'Codice Tributo',
            'AMMONTARE LORDO CORRISPOSTO',
            'TOTALE SOMME NON SOGGETTE A RITENUTA D\'ACCONTO',
            'IMPONIBILE IRPEF', '% RITENUTA D\'ACCONTO',
            'IMPORTO RITENUTA D\'ACCONTO', 'Codice Somme non soggette a RdA',
            '% COMPENSO SOGGETTA A RITENUTA D\'ACCONTO'
        ]
        safe = pd.DataFrame()
        for col in required_columns:
            if col in table_data.columns:
                safe[col] = table_data[col]
            else:
                sim = [c for c in table_data.columns if col.replace("'", "") in c.replace("'", "")]
                safe[col] = table_data[sim[0]] if sim else ''

        # -------------------------------------------------------------------
        #  ANAGRAFICHE
        # -------------------------------------------------------------------
        anag_file = io.BytesIO()
        mask = safe['C. Fisc'].notna() & (safe['C. Fisc'] != '')
        anagrafiche_data = safe[mask].copy() if mask.any() else pd.DataFrame({
            'C. Fisc': ['ESEMPIO12345ABCDE'],
            'P. Iva': ['IT12345678901'],
            'Cognome / Denominazione Percipiente': ['ESEMPIO COGNOME'],
            'Nome Percipiente': ['ESEMPIO NOME']
        })

        # Usa l'anno corrente per Anno Certificazione Unica
        anno_corrente = str(datetime.datetime.now().year)
        logging.info(f"Anno corrente: {anno_corrente}")

        anag_out = pd.DataFrame()
        # Assegna esplicitamente l'anno corrente
        anag_out['Anno Certificazione Unica'] = pd.Series([anno_corrente] * len(anagrafiche_data))
        anag_out['Codice Fiscale Sostituto'] = ''
        anag_out['Tipo soggetto'] = ''
        anag_out['Codice Fiscale Percipiente'] = anagrafiche_data['C. Fisc'].apply(format_text_for_excel)
        anag_out['Partita IVA del Percipiente'] = anagrafiche_data['P. Iva'].apply(clean_piva)
        anag_out['Cognome / Denominazione Percipiente'] = anagrafiche_data['Cognome / Denominazione Percipiente']
        anag_out['Nome Percipiente'] = anagrafiche_data['Nome Percipiente']

        # colonne anagrafiche aggiuntive vuote (dtype object)
        empty_cols_anag = [
            'Sesso Percipiente',
            'Data di nascita Percipiente (GGMMAAAA)',
            'Comune o Stato estero di nascita Percipiente',
            'Provincia del comune di nascita Percipiente',
            'Domicilio al 01/01/2024 Comune', 'Domicilio 01/01/2024 Provincia',
            'Domicilio al 01/01/2024 Codice IRPEF', 'Fusione Comune',
            'Domcilio al 01/01/2025 Comune', 'Domcilio al 01/01/2025 Provincia',
            'Domcilio al 01/01/2025 Codice IRPEF Comune', 'Fusione Comune 2',
            'Codice fiscale Rappresentante', 'Codice identificazione fiscale estero',
            'Località di residenza estera', 'Luogo di attinenza',
            'Indirizzo di residenza estera', 'Non residenti Schumacker',
            'Codice stato estero', 'Frontaliere',
            'Residenza o Sede legale - Indirizzo', 'Residenza o Sede legale - Stato',
            'Residenza o Sede legale - Comune', 'Residenza o Sede legale - Provincia',
            'Residenza o Sede legale - CAP', 'Residenza o Sede legale - Email',
            'Residenza o Sede legale - PEC', 'Domicilio fiscale - Indirizzo',
            'Domicilio fiscale - Stato', 'Domicilio fiscale - Comune',
            'Domicilio fiscale - Provincia', 'Domicilio fiscale - CAP',
            'Domicilio fiscale - Email', 'Domicilio fiscale - PEC',
            'Docenti ricercatori e impatriati Codice stato estero'
        ]
        for col in empty_cols_anag:
            anag_out[col] = ''

        # riempi info CF
        for idx, cf in enumerate(anagrafiche_data['C. Fisc']):
            if isinstance(cf, str) and len(cf) == 16:
                info = extract_info_from_cf(cf, db_password)
                if info:
                    anag_out.at[idx, 'Sesso Percipiente'] = info['sesso']
                    anag_out.at[idx, 'Data di nascita Percipiente (GGMMAAAA)'] = info['data_nascita']
                    anag_out.at[idx, 'Comune o Stato estero di nascita Percipiente'] = info['comune_nascita']
                    anag_out.at[idx, 'Provincia del comune di nascita Percipiente'] = info['provincia_nascita']

        anag_out = anag_out.drop_duplicates()

        with pd.ExcelWriter(anag_file, engine='xlsxwriter') as w:
            anag_out.to_excel(w, sheet_name='Anagrafiche', index=False)
            workbook = w.book
            worksheet = w.sheets['Anagrafiche']
            
            # Formato per preservare gli zeri iniziali
            text_format = workbook.add_format({'num_format': '@'})
            
            # Applica formato testo alle colonne Codice Fiscale e Partita IVA
            cf_col_idx = anag_out.columns.get_loc('Codice Fiscale Percipiente') + 1  # +1 perché Excel usa indici 1-based
            piva_col_idx = anag_out.columns.get_loc('Partita IVA del Percipiente') + 1
            worksheet.set_column(cf_col_idx, cf_col_idx, None, text_format)
            worksheet.set_column(piva_col_idx, piva_col_idx, None, text_format)
            
        anag_file.seek(0)
        output1_base64 = base64.b64encode(anag_file.getvalue()).decode()

        # -------------------------------------------------------------------
        #  COMPENSI
        # -------------------------------------------------------------------
        comp_file = io.BytesIO()
        comp_data = safe[mask].copy() if mask.any() else anagrafiche_data

        # Definizione dell'ordine completo delle colonne secondo il documento fornito
        all_columns = [
            'Anno Certificazione Unica',
            'Codice fiscale Sostituto',
            'Codice fiscale Percipiente',
            'Partita IVA Percipiente',
            'Codice identificazione fiscale estero',
            'Cognome / Denominazione Percipiente',
            'Nome Percipiente',
            'Categoria particolare',
            'Eventi eccezionali',
            'Esclusione dalla precompilata',
            'Tipo operazione',
            'Protocollo comunicazione da sostituire o da annullare - identificativo invio',
            'Protocollo documento da sostituire o da annullare - progressivo attribuito alla C.U. originaria',
            'CU AUTONOMO Tipologia reddituale - Causale (punto 1)',
            'Anno di competenza compenso',
            'Tipo partitario',
            'Percipiente / partitario',
            'Causale',
            'Codice tributo',
            'Aliquota ritenuta',
            'Data documento (GGMMAAAA)',
            'Numero fattura / documento',
            'Tipo documento',
            'Data di pagamento (GGMMAAAA)',
            'Tipo operazione ritenute',
            'CU AUTONOMO Dati fiscali - Anno (punto 2)',
            'CU AUTONOMO Dati fiscali - Anticipazione (punto 3)',
            'CU AUTONOMO   Dati fiscali - Ammontare lordo corrisposto (punto 4)',
            'CU AUTONOMO Dati fiscali - Somme non soggette a ritenuta regime convenzionale (punto 5)',
            'Percentuale imponibilita\'',
            'Aliquota ritenuta d\'acconto',
            'Aliquota ritenuta d\'imposta',
            'Altre somme non soggette - Deduzioni / provvigioni',
            'Altre somme non soggette - Somme non soggette',
            'Altre somme non soggette - Somme che non concorrono al reddito',
            'CU AUTONOMO Dati fiscali - Codice (punto 6)',
            'CU AUTONOMO Dati fiscali - Altre somme non soggette a ritenuta (punto 7)',
            'CU AUTONOMO Dati fiscali - Imponibile (punto 8)',
            'CU AUTONOMO Dati fiscali - Ritenute a titolo d\'acconto (punto 9)',
            'CU AUTONOMO  Dati fiscali - Ritenute a titolo d\'imposta (punto 10)',
            'CU AUTONOMO Dati fiscali - Ritenute sospese (punto 11)',
            'CU AUTONOMO  Dati fiscali - Addizionale regionale a titolo d\'acconto (punto 12)',
            'CU AUTONOMO   Dati fiscali - Addizionale regionale a titolo d\'imposta (punto 13)',
            'CU AUTONOMO   Dati fiscali - Addizionale regionale sospesa (punto 14)',
            'CU AUTONOMO   Dati fiscali - Addizionale comunale a titolo d\'acconto (punto 15)',
            'CU AUTONOMO   Dati fiscali - Addizionale comunale a titolo d\'imposta (punto 16)',
            'CU AUTONOMO   Dati fiscali - Addizionale comunale sospesa (punto 17)',
            'CU AUTONOMO    Dati fiscali - Imponibile anni precedenti (punto 18)',
            'CU AUTONOMO Dati fiscali - Ritenute operate anni precedenti (punto 19)',
            'CU AUTONOMO Dati fiscali - Spese rimborsate (punto 20)',
            'CU AUTONOMO Dati fiscali - Somme restituite al netto della ritenuta subita (punto 22)',
            'CU AUTONOMO Dati previdenziali - Codice Fiscale Ente Previdenziale (punto 29)',
            'CU AUTONOMO Dati previdenziali - Denominazione Ente Previdenziale (punto 30)',
            'CU AUTONOMO Dati previdenziali - Tipo di rapporto (punto 31)',
            'CU AUTONOMO Dati previdenziali - Codice azienda (punto 32)',
            'CU AUTONOMO Dati previdenziali - Categoria (punto 33)',
            'CU AUTONOMO Dati previdenziali - Contributi previdenziali a carico del soggetto erogante (punto 34)',
            'CU AUTONOMO Dati previdenziali - Contributi previdenziali a carico del percipiente (punto 35)',
            'CU AUTONOMO Dati previdenziali - Altri contibuti (punto 36)',
            'CU AUTONOMO Dati previdenziali - Importo altri contributi (punto 37)',
            'CU AUTONOMO Dati previdenziali - Contributi dovuti (punto 38)',
            'CU AUTONOMO Dati previdenziali - Contributi versati (punto 39)',
            'CU AUTONOMO Dati previdenziali - Importo franchigia utilizzata (punto 40)',
            'CU AUTONOMO Fallimento e liquidazione coatta amministrativa - Somme corrisposte prima della data di fallimento (punto 41)',
            'CU AUTONOMO Fallimento e liquidazione coatta amministrativa - Somme corrisposte dal curatore /commissario (punto 42)',
            'CU AUTONOMO Redditi erogati da altri soggetti - Codice fisale (punto 52)',
            'CU AUTONOMO Redditi erogati da altri soggetti - Imponibile (punto 53)',
            'CU AUTONOMO Redditi erogati da altri soggetti - Ritenute a titolo d\'acconto (punto 54)',
            'CU AUTONOMO Redditi erogati da altri soggetti - Ritenute a titolo d\'imposta (punto 55)',
            'CU AUTONOMO Redditi erogati da altri soggetti - Ritenute sospese (punto 56)',
            'CU AUTONOMO Redditi erogati da altri soggetti - Add. Regionale a titolo d\'acconto (punto 57)',
            'CU AUTONOMO Redditi erogati da altri soggetti - Add. Regionale a titolo d\'imposta (punto 58)',
            'CU AUTONOMO Redditi erogati da altri soggetti - Add. Regionale sospesa (punto 59)',
            'CU AUTONOMO Redditi erogati da altri soggetti - Add. Comunale a titolo d\'acconto (punto 60)',
            'CU AUTONOMO Redditi erogati da altri soggetti - Add. Comunale a titolo d\'imposta (punto 61)',
            'CU AUTONOMO Redditi erogati da altri soggetti - Add. Comunale sospesa (punto 62)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati - Codice Fiscale PPAA/Azienda (punto 52)',
            'CU DIPENDENTE   INPS Gestione separata Parasubordinati - Tipo rapporto (punto 51)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati - Compensi corrisposti al parasubordinato (punto 45)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati - Contributi dovuti (punto 46)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati - Contributi a carico del lavoratore (punto 47)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati - Contributi versati (punto 48)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati - Tutti (punto 49)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati - Gennaio (punto 50)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati - Febbraio (punto 50)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati - Marzo (punto 50)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati - Aprile (punto 50)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati - Maggio (punto 50)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati - Giugno (punto 50)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati - Luglio (punto 50)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati - Agosto (punto 50)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati - Settembre (punto 50)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati - Ottobre (punto 50)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati - Novembre (punto 50)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati - Dicembre (punto 50)',
            'CU DIPENDENTE       INPS Gestione separata Parasubordinati Sportivi dilettantistici - Tipo rapporto (punto 61)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati Sportivi dilettantistici - Altro tipo rapporto (punto 62)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati Sportivi dilettantistici - Compensi totali parasubordinati sportivi e assimilati (punto 53)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati Sportivi dilettantistici - Imponibile contributivo (punto 54)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati Sportivi dilettantistici - Imponibile IVS (punto 55)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati Sportivi dilettantistici - Contributi dovuti (punto 56)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati Sportivi dilettantistici - Contributi a carico del parasubordinato (punto 57)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati Sportivi dilettantistici - Contributi versati (punto 58)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati Sportivi dilettantistici - Tutti (punto 59)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati Sportivi dilettantistici - Gennaio (punto 60)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati Sportivi dilettantistici - Febbraio (punto 60)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati Sportivi dilettantistici- Marzo (punto 60)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati Sportivi dilettantistici- Aprile (punto 60)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati Sportivi dilettantistici - Maggio (punto 60)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati Sportivi dilettantistici- Giugno (punto 60)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati Sportivi dilettantistici - Luglio (punto 60)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati Sportivi dilettantistici- Agosto (punto 60)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati Sportivi dilettantistici- Settembre (punto 60)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati Sportivi dilettantistici- Ottobre (punto 60)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati Sportivi dilettantistici- Novembre (punto 60)',
            'CU DIPENDENTE INPS Gestione separata Parasubordinati Sportivi dilettantistici- Dicembre (punto 60)'
        ]

        # Inizializza il DataFrame con tutte le colonne vuote
        comp_out = pd.DataFrame(columns=all_columns)
        num_rows = len(comp_data)
        
        # Popola le colonne con i dati disponibili, assicurandoci che tutte abbiano la lunghezza corretta
        comp_out['Anno Certificazione Unica'] = pd.Series([anno_corrente] * num_rows)
        comp_out['Codice fiscale Sostituto'] = pd.Series([''] * num_rows)
        comp_out['Codice fiscale Percipiente'] = comp_data['C. Fisc'].apply(format_text_for_excel)
        comp_out['Partita IVA Percipiente'] = comp_data['P. Iva'].apply(clean_piva)
        comp_out['Codice identificazione fiscale estero'] = pd.Series([''] * num_rows)
        comp_out['Cognome / Denominazione Percipiente'] = comp_data['Cognome / Denominazione Percipiente']
        comp_out['Nome Percipiente'] = comp_data['Nome Percipiente']
        comp_out['CU AUTONOMO Tipologia reddituale - Causale (punto 1)'] = comp_data.get('Causale CU', '')
        comp_out['Anno di competenza compenso'] = pd.Series([str(datetime.datetime.now().year - 1)] * num_rows)
        comp_out['Codice tributo'] = comp_data.get('Codice Tributo', '')

        # Gestione aliquota ritenuta
        if '% RITENUTA D\'ACCONTO' in comp_data.columns:
            aliq = comp_data['% RITENUTA D\'ACCONTO'].apply(_percent_times_100)
            comp_out['Aliquota ritenuta'] = aliq
            comp_out['Aliquota ritenuta d\'acconto'] = aliq

        # Percentuale compenso soggetta a ritenuta (non nella lista delle colonne finale, ma presente nell'input)
        if '% COMPENSO SOGGETTA A RITENUTA D\'ACCONTO' in comp_data.columns:
            perc_comp = comp_data['% COMPENSO SOGGETTA A RITENUTA D\'ACCONTO'].apply(_percent_times_100)
            # La usiamo per calcolare eventualmente altri campi se necessario

        # Ammontare lordo corrisposto
        comp_out['CU AUTONOMO   Dati fiscali - Ammontare lordo corrisposto (punto 4)'] = (
            comp_data['AMMONTARE LORDO CORRISPOSTO'].apply(_to_float)
            if 'AMMONTARE LORDO CORRISPOSTO' in comp_data.columns else ''
        )

        # Codice somme non soggette
        comp_out['CU AUTONOMO Dati fiscali - Codice (punto 6)'] = comp_data.get('Codice Somme non soggette a RdA', '')

        # Altre somme non soggette a ritenuta
        comp_out['CU AUTONOMO Dati fiscali - Altre somme non soggette a ritenuta (punto 7)'] = (
            comp_data.get('TOTALE SOMME NON SOGGETTE A RITENUTA D\'ACCONTO', '').apply(_to_float)
            if 'TOTALE SOMME NON SOGGETTE A RITENUTA D\'ACCONTO' in comp_data.columns else ''
        )

        # Imponibile
        comp_out['CU AUTONOMO Dati fiscali - Imponibile (punto 8)'] = (
            comp_data.get('IMPONIBILE IRPEF', '').apply(_to_float)
            if 'IMPONIBILE IRPEF' in comp_data.columns else ''
        )

        # Ritenute a titolo d'acconto
        comp_out['CU AUTONOMO Dati fiscali - Ritenute a titolo d\'acconto (punto 9)'] = (
            comp_data.get('IMPORTO RITENUTA D\'ACCONTO', '').apply(_to_float)
            if 'IMPORTO RITENUTA D\'ACCONTO' in comp_data.columns else ''
        )

        # Configura il writer per preservare il formato testo con zeri iniziali
        with pd.ExcelWriter(comp_file, engine='xlsxwriter') as w:
            comp_out.to_excel(w, sheet_name='Compensi', index=False)
            workbook = w.book
            worksheet = w.sheets['Compensi']
            
            # Formato per preservare gli zeri iniziali
            text_format = workbook.add_format({'num_format': '@'})
            
            # Applica formato testo alle colonne Codice Fiscale e Partita IVA
            cf_col_idx = 2  # La colonna 'Codice fiscale Percipiente' (indice 0-based + 1 per Excel)
            piva_col_idx = 3  # La colonna 'Partita IVA Percipiente' (indice 0-based + 1 per Excel)
            worksheet.set_column(cf_col_idx, cf_col_idx, None, text_format)
            worksheet.set_column(piva_col_idx, piva_col_idx, None, text_format)
            
        comp_file.seek(0)
        output2_base64 = base64.b64encode(comp_file.getvalue()).decode()

        return func.HttpResponse(
            json.dumps({
                "output1Base64": output1_base64,
                "output2Base64": output2_base64
            }),
            mimetype="application/json"
        )

    except Exception as e:
        logging.error(f'Errore: {e}')
        logging.error(traceback.format_exc())
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )